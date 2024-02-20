# Migrate heliogaia.de calculations from Excel to Python

# this is an extract for overview, but
# the main development is done in Quarto in notebooks/Excel2Python_Migrator.qmd

# Es sollen zunächst die Heiz-Kosten pro Monat und Kopf mit einem 
# Heliogaia-System für die Gemeinde Röbel nachvollzogen werden.

# Vorerst nur:
# Zielszenario, ohne Blockheizkraftwerk (BHKW), Menschen leben in sanierten Gebäuden mit durchschnittlich 80 kWh/a/m²
# 
# Erweiterungsmöglichkeit:
# Übergangsszenario, mit BHKW, nicht alle Einwohner leben in sanierten Gebäuden, drei Fernwärmeleitungen nötig

# first empty the Python environment
Z_my_previous_python_variables =\ 
  [k for k in globals().keys() if not (k.startswith("__") | (k == "r")) ]
print(Z_my_previous_python_variables)

for key in Z_my_previous_python_variables:
  exec('del('+key+')')


import pandas as pd
from openpyxl import load_workbook
import webbrowser
import re
from re import findall
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from datetime import datetime
import pickle


from openpyxl.utils import get_column_letter
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.utils.cell.html?highlight=openpyxl.utils.cell.get_column_letter#openpyxl.utils.cell.get_column_letter

from openpyxl.utils import column_index_from_string
#https://openpyxl.readthedocs.io/en/stable/api/openpyxl.utils.cell.html
#```


### Beispielkommune Roebel
#### Datei jahreslauf_roebel.xlsx' von Heliogaia.de geladen ca. am 12.01.2024. 
##### Blatt e

sheet_e = pd.read_excel(
  'notebooks/_base/jahreslauf_roebel.xlsx'
  , sheet_name='e', header = 1)
print(sheet_e.shape)

kurven = sheet_e.iloc[1:732,:18]; kurven.head()
einheiten = sheet_e.iloc[0,:18]
original = kurven.copy()

# pd.DataFrame(list(original)).to_clipboard(header=False, index=False)
# pd.DataFrame(list(einheiten)).to_clipboard(header=False, index=False)

#make_var_name
def make_var_name(input_string):
  """
  Generate a python object name from a variable description
  """
  inter = input_string
  inter = re.sub('\.', '', inter)
  inter = re.sub(', entspricht', '', inter)
  inter = re.sub('[\s:\(\),;-]', '_', inter)
  inter = re.sub('ä', 'ae', inter)
  inter = re.sub('ö', 'oe', inter)
  inter = re.sub('ü', 'ue', inter)
  inter = re.sub('Ä', 'Ae', inter)
  inter = re.sub('Ö', 'Oe', inter)
  inter = re.sub('Ü', 'Ue', inter)
  inter = re.sub('ß', 'ss', inter)
  inter = re.sub('&', 'u' , inter)
  inter = re.sub('\+', '_u_' , inter)
  inter = re.sub('/', '_pro_', inter)
  inter = re.sub('²', '2', inter)
  inter = re.sub('³', '3', inter)
  inter = re.sub('%', 'Prozent', inter)
  inter = re.sub('°', '_Grad_', inter)
  inter = re.sub('=', '_gleich_', inter)
  inter = re.sub('€', '_Euro_', inter)
  inter = re.sub('±', '_plusminus_', inter)
  return inter

#test cases
make_var_name('hi;')

new_column_names = [make_var_name(col) for col in list(original)]

kurven.columns = new_column_names
new_column_names
kurven


titles = [
  'Datum'
  ,'Endenergie (EE) Verbrauch  für Heizung & Warmwasser [kWh/d/Kopf]'
  ,'Endenergie Direktbezug aus Kollektoren, am Saisonspeicher vorbei [kWh/d/Kopf]'
  ,'beim Verbraucher verfügbare EE aus innerörtlichen Kollektoren [kWh/d/Kopf]'
  ,'beim Verbraucher verfügbare EE aus  externen   Kollektoren [kWh/d/Kopf]'
  ,'sofort verbrauchter Anteil des innerörtlichen Kollektorgewinns  [%]'
  ,'sofort verbrauchter Anteil des externen Kollektorgewinns, als Fernwärme  [%]'
  ,'am Saisonspeicher verfügbare Wärme aus innerörtlichen Kollektoren [kWh/d/Kopf]'
  ,'am Saisonspeicher verfügbare Wärme aus externen Kollektoren [kWh/d/Kopf]'
  ,'Deckungsgrad, allein aus innerörtlichen Kollektoren [%]'
  ,'Fernwärmebezug aus externem Kollektorfeld [kWh/d/Kopf]'
  ,'Fernwärmebezug aus Saisonspeicher [kWh/d/Kopf]'
  ,'Saisonspeicher Belastung [kWh/d/Kopf]'
  ,'Fernwärme Bezug [kWh/d/Kopf]'
  ,'Speicher laden [kWh/d/Kopf]'
  ,'Speicher Iinhalt [kWh/Kopf]'
  ,'Speicher Temperatur [°C]'
]


fig, axs = plt.subplots(nrows=17, figsize=[13,80])
for i, col in enumerate(new_column_names[1:]):
  sns.scatterplot(x='Tag', y=col, data=kurven, ax=axs[i])
  axs[i].set_title(titles[i])
fig.tight_layout()
#```


##### Blatt t

#```{python dedup}
list_of_steps = [] 


step16 = pd.DataFrame({'sheetcell': [
  'u!B12',
  'u!B11',
  'h!G16',
  'h!D16'
  ]})
step16.loc[:, 'step'] = 16
list_of_steps = list_of_steps + [step16]



step15 = pd.DataFrame({'sheetcell': [
  'u!K23',
  'u!K24',
  'h!P21', # 2024-01-23--23:56:20
  'h!P22',
  'h!P23',
  'h!P24',
  'h!P25',
  'h!P26',
  'h!P27',
  'h!P28',
  'h!P29',
  'h!P30',
  'h!P31'
  # 's!Q2:Q31'
  ]})
step15.loc[:, 'step'] = 15
list_of_steps = list_of_steps + [step15]



step14 = pd.DataFrame({'sheetcell': [
  'u!L23',
  'u!L24',
  's!F20',
  's!F24',
  's!F12',
  's!F10',
  's!F22',
  's!F14',
  's!F7',
  's!F8', #added manually 2024-01-23--18:49:11
  's!F2',
  # added 2024-01-23--23:51:51 :
  'h!Q21',
  'h!Q22',
  'h!Q23',
  'h!Q24',
  'h!Q25',
  'h!Q26',
  'h!Q27',
  'h!Q28',
  'h!Q29',
  'h!Q30',
  'h!Q31',
  's!Q32'
  ]})
step14.loc[:, 'step'] = 14
list_of_steps = list_of_steps + [step14]



step13 = pd.DataFrame({'sheetcell': [
  't!D52',
  't!D56',
  'u!L25',
  's!G20',
  's!G24',
  's!G12',
  's!G10',
  's!G22',
  's!G14',
  's!G7',
  's!G8', #added manually 2024-01-23--18:49:02
# start added 2024-01-23--23:46:03
  't!D61',
  't!D58',
  't!D57',
  't!D59',
  't!D53',
  't!D54',
  't!D55',
  'h!Q32',
  's!O35',
  #end added 2024-01-23--23:46:23
  's!G2'
  ]})
step13.loc[:, 'step'] = 13
list_of_steps = list_of_steps + [step13]


step12 = pd.DataFrame({'sheetcell': [
  't!D109', #für hohe Genauigkeitsansprüche bei jeder Parameteränderung extern mit zylindermodell007.ods neu berechnen
  't!D110',
  't!D111',
  't!D112',
  't!D113',
  't!D131',
  # 's!Q2:Q31',
  'h!D11',
  'h!D13',
  'h!D9',
  'h!D8',
  'h!D12',
  'h!D10',
  'h!D6',
  'h!D7', #added manually 2024-01-23--18:32:30
 #start added 2024-01-23--23:37:50
  't!D137',
  's!O36',
  # 's!D2:D31' #end
  'h!D5'
  ]})
step12.loc[:, 'step'] = 12
list_of_steps = list_of_steps + [step12]



step11 = pd.DataFrame({'sheetcell': [
  't!D114',
  't!D145',
  # 's!Q32',
  'h!D21',
  'h!D22',
  'h!D23',
  'h!D24',
  'h!D25',
  'h!D26',
  'h!D27',
  'h!D28',
  'h!D29',
  'h!D30',
  'h!D31',
  'u!G24',
  'u!G23',
  'u!B6',  #added 2024-01-23--23:34:53
  's!D32'
  ]})
step11.loc[:, 'step'] = 11
list_of_steps = list_of_steps + [step11]



step10 = pd.DataFrame({'sheetcell': [
  't!D153',
  't!D154',
  't!D31',
  'h!F21',
  'h!F22',
  'h!F23',
  'h!F24',
  'h!F25',
  'h!F26',
  'h!F27',
  'h!F28',
  'h!F29',
  'h!F30',
  'h!F31',
  'u!H24',
  'u!H23',
  'h!I21',
  'h!I22',
  'h!I23',
  'h!I24',
  'h!I25',
  'h!I26',
  'h!I27',
  'h!I28',
  'h!I29',
  'h!I30',
  'h!I31',
  'u!B9',  #start added  2024-01-23--23:29:31
  'u!D16',
  's!I35',
  'u!F24',
  'u!F23',
  'u!B4'    #end
  ]})
step10.loc[:, 'step'] =10
list_of_steps = list_of_steps + [step10]




step09 = pd.DataFrame({'sheetcell': [
  'e!V25',
  't!D155',
  'u!G16',
  'u!E16',
  'u!B7',
  'u!G17',
  'u!E17',
  't!D118',
  'h!G21',
  'h!G22',
  'h!G23',
  'h!G24',
  'h!G25',
  'h!G26',
  'h!G27',
  'h!G28',
  'h!G29',
  'h!G30',
  'h!G31',
  'u!I24',
  'u!I23',
  'h!K21',
  'h!K22',
  'h!K23',
  'h!K24',
  'h!K25',
  'h!K26',
  'h!K27',
  'h!K28',
  'h!K29',
  'h!K30',
  'h!K31',
  'e!V27',
  'u!C16', #start  manually  2024-01-23--23:24:41
  'e!V28',
  'e!V29',
  'e!V11' #end
  ]})
step09.loc[:, 'step'] = 9
list_of_steps = list_of_steps + [step09]

#start


step08 = pd.DataFrame({'sheetcell': [
  't!D99',
  't!D28',
  't!D60',
  't!D159',
  't!D66',
  'u!H16',
  'u!F16',
  'u!B8',
  'u!H17',
  'u!F17',
  't!D119',
  'h!J21',
  'h!H21',
  'h!D17',
  'h!J22',
  'h!H22',
  'h!J23',
  'h!H23',
  'h!J24',
  'h!H24',
  'h!J25',
  'h!H25',
  'h!J26',
  'h!H26',
  'h!J27',
  'h!H27',
  'h!J28',
  'h!H28',
  'h!J29',
  'h!H29',
  'h!J30',
  'h!H30',
  'h!J31',
  'h!H31',
  't!D7',
  's!O32',
  'u!I25',
  'h!K32',
  'e!E4',
  'e!F4',
  'e!V30',
  'e!A4',
  'e!V12',
  'e!V13'
  ]}) 
step08.loc[:, 'step'] = 8
list_of_steps = list_of_steps + [step08]



step07 = pd.DataFrame({'sheetcell': [
  't!D103',
  't!D64',
  'u!I16',
  'u!I17',
  't!D121',
  'h!L21',
  'h!L22',
  'h!L23',
  'h!L24',
  'h!L25',
  'h!L26',
  'h!L27',
  'h!L28',
  'h!L29',
  'h!L30',
  'h!L31',
  'e!V9',
  'e!V10',
  'e!V14',
  'e!V17',
  't!D90',
  't!D91',
  's!O34',
  'h!E21',
  'h!E22',
  'h!E23',
  'h!E24',
  'h!E25',
  'h!E26',
  'h!E27',
  'h!E28',
  'h!E29',
  'h!E30',
  'h!E31',
  't!D129',
  't!D135',
  'e!V6',
  'e!V7',
  'e!D4', #'e!D4:D734' #hier soll vorerst eine Zelle reichen, um die df Formeln zu bekommen
  'e!C4', #'e!C4:C734' #dann muss über die DataFrame Spalte summiert werden
  'e!V15',
  'e!V18',
  'e!V33'
  ]}) 
step07.loc[:, 'step'] = 7
list_of_steps = list_of_steps + [step07]



step06 = pd.DataFrame({'sheetcell': [
  't!D106',
  't!D62',
  't!D68',
  't!D67',
  't!D63',
  't!D65',
  't!D70',
  't!D71',
  't!D105',
  't!D69',
  't!D72',
  't!D115',
  't!D73',
  'u!I18',
  't!D122',
  'h!L32',
  'e!V16',
  't!D40',
  't!D12',
  't!D13',
  't!D17',
  't!D39',
  't!D20',
  't!D92',
  't!D44',
  't!D120',
  's!I32',
  'h!E32',
  't!D139',
  't!D34',
  't!D101',
  'e!V5',
  'e!V8',
  'e!D736',
  'e!C736',
  't!D18',
  't!D21',
  't!D29'
  ]})
step06.loc[:, 'step'] = 6
list_of_steps = list_of_steps + [step06]



step05 = pd.DataFrame({'sheetcell': [
  't!D166', # step05
  't!D167',
  't!D168',
  't!D169',
  't!D170',
  't!D130',
  't!D80',
  't!D136',
  't!D41',
  't!D42',
  't!D19',
  't!D46',
  't!D45',
  't!D48',
  't!D49',
  't!D47',
  't!D93',
  't!D76',
  't!D127',
  't!D77',
  't!D117',
  't!D75',
  't!D133',
  't!F41',
  't!F45',
  't!F49',
  't!F77',
  't!F75',
  't!D143',
  't!D36',
  't!D83',
  'e!V20',
  't!D8',
  't!D11',
  'e!D737',
  'e!V24',
  't!D89', #added after check
  't!F80' #added after check
  ]})
step05.loc[:, 'step'] = 5
list_of_steps = list_of_steps + [step05]

step04 = pd.DataFrame({'sheetcell': [
  't!D171', #step04
  't!D172',
  't!D173',
  't!D174',
  't!D175',
  't!D176',
  't!D177',
  't!D178',
  't!D179',
  't!F171',
  't!F172',
  't!F173',
  't!F174',
  't!F175',
  't!F176',
  't!F177',
  't!F178',
  't!F179',
  't!D149',
  't!D162',
  't!D87',
  't!D78',
  't!D164',
  't!D85',
  't!D81',
  't!D33',
  't!D23',
  't!D32',
  't!D35',
  't!D27'
  ]})
step04.loc[:, 'step'] = 4
list_of_steps = list_of_steps + [step04]

step03 = pd.DataFrame({'sheetcell': [
  't!D189', # step03
  't!D190',
  't!D191',
  't!D192',
  't!D193',
  't!D194',
  't!D195',
  't!D196',
  't!D197',
  't!D183',
  't!D186',
  't!D187',
  't!D86',
  't!F81',
  'e!V4',
  't!D100',
  't!D98',
  't!D84',
  't!D161',
  't!D163'
  ]})
step03.loc[:, 'step'] = 3
list_of_steps = list_of_steps + [step03]

step02 = pd.DataFrame({'sheetcell': [
  't!D199', #step02
  't!D37',
  't!D202',
  't!D205',
  't!D206',
  't!D180',
  't!F180',
  't!E3',
  't!D184',
  't!D185'
  ]})
step02.loc[:, 'step'] = 2
list_of_steps = list_of_steps + [step02]

step01 = pd.DataFrame({'sheetcell': [
  't!D200', #step01
  't!D207',
  't!D198',
  't!D203',
  't!D204'
  ]})
step01.loc[:, 'step'] = 1  
list_of_steps = list_of_steps + [step01]


dedup = pd.concat(list_of_steps, axis=0)
dedup = dedup.reset_index(drop=True)

# remove dollar sign
dedup.loc[:, 'sheetcell'] = dedup.loc[:, 'sheetcell'].str.replace(r'$', '', regex=True)

#keep for a later step
pure_dedup = dedup.copy()

## merge old dna
old_dna = pd.read_csv('notebooks/part_results/workbook_dna.csv')
dedup = dedup.merge(old_dna, on='sheetcell', how='left')
dedup = dedup.loc[:, ['sheetcell', 'step_x', 'sheetformula']]
dedup.columns = ['sheetcell', 'step', 'sheetformula'] #rename step_x
old_dna = dedup.copy() #important, otherwise old_dna.update below deletes floor

if dedup.loc[:,'sheetcell'].duplicated().sum() > 0:
  print('Doppelt sind: ',
    dedup.loc[dedup.duplicated(), :]
  )
else: print('OK, no duplicates!')

#```




###### dedup_check --

#```{python formulator - separate cells and sheets and delete dollar sign}
formulator = dedup.copy()

# extract sheet
formulator.loc[:, 'sheet'] = \
  formulator.loc[:, 'sheetcell']\
  .str.extract(r'^(\w{1,})\!', expand=False) #do not expand Series to DataFrame
#check
formulator['sheet'].value_counts() #ok
# formulator.loc[35, :]

# remove dollar sign
formulator.loc[:, 'sheetcell'] = \
  formulator.loc[:, 'sheetcell']\
  .str.replace(r'$', '', regex=True)
  
# formulator.loc[:, 'sheetcell'] = formulator.loc[:, 'cell']

#remove sheet prefix
# formulator.loc[mask_other_sheets, 'cell'] = \
formulator.loc[:, 'cell'] = \
  formulator.loc[:, 'sheetcell']\
  .str.replace(r'^(\w{1,})\!', '', regex=True)
#check
# formulator.loc[35, :] #ok

# extract column
formulator.loc[:, 'column'] = \
  formulator.loc[:, 'cell']\
  .str.extract(r'^([A-Z]{1,})\d{1,}', expand=False)
# formulator['column'].value_counts() #ok
  
# extract row
formulator.loc[:, 'row'] = \
  formulator.loc[:, 'cell']\
  .str.extract(r'^[A-Z]{1,}(\d{1,})', expand=False)
formulator['row'].value_counts() #ok

#```


#```{python formulator  add formulas}
# webbrowser.open('https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.html')

# load_workbook
wbf_workbook = load_workbook('notebooks/_base/jahreslauf_roebel.xlsx') #workbook containing formulas incl constants
wbv_workbook_values = load_workbook('notebooks/_base/jahreslauf_roebel.xlsx', data_only=True)

# add formulas
for i in formulator.index:
  # treat special cases
  set_as_input = ['e!V4', 's!I32', 's!O32', 't!D31'] #set an input value instead of looking at calculation
  if formulator.loc[i, 'sheetcell'] in set_as_input:
    formulator.loc[i, 'formula'] = wbv_workbook_values[formulator.loc[i,'sheet']][formulator.loc[i,'cell']].internal_value
  else:
    formulator.loc[i, 'formula'] =\
    wbf_workbook[formulator.loc[i,'sheet']][formulator.loc[i,'cell']].internal_value

# using strings above here will silence the FutureWarning, 
# but alter the subsequent calculations, which would have to be 
# tested thoroughly 2024-01-25

# formulator.iloc[75:, :]

# r.View(formulator, 'formulator')

#```

###### regex

#```{python formulator - regex for }

# from re import findall
# todo replace {0,1} where possible in the below regex
regex_cell_extraction = r'\b([A-Za-z]{0,}\!{0,1}\${0,1}[A-Z]{1,2}\${0,1}\d{1,}\:{0,1}[A-Za-z]{0,}\!{0,1}\${0,1}[A-Z]{0,2}\${0,1}\d{0,})\b'
#checks
# formulator.loc[[0, 9, 35, 10, 67, 80, 97],'formula'] #example cases
# formulator.loc[:,'formula']\
# formulator.loc[[0, 9, 35, 10, 67, 80, 97],'formula']\
  # .str.findall(regex_cell_extraction)

#```

#```{python formulator  whole check}
formulator.loc[:,'next_cells'] = \
  formulator.loc[:,'formula']\
    .str.findall(regex_cell_extraction)

# # refurbish with sheet
# mask_other_sheets = (~formulator.loc[:, 'sheet'] == current_sheet) & (~formulator.loc[:, 'formula'].str.contains(r'\!'))
# formulator.loc[mask_other_sheets, 'next_cells'] = formulator.loc[mask_other_sheets, 'next_cells'].str.replace(r'[A-Z]{1,2}',)

#check    
# formulator.loc[[0, 9, 35, 10, 80, 97],:] #example cases
# # overlong cases
# formulator.loc[82,'formula']
# formulator.loc[82,'next_cells']
# formulator.loc[11,'formula']
# formulator.loc[11,'next_cells']
# formulator.head(6)
# r.View(formulator, 'formulator')
#```

###### from_step corner

#```{python clipboarder}
from_step=16
mask_for_clipboard = (formulator.loc[:,'step']==from_step)\
  &(~formulator.loc[:,'next_cells'].isna())
sum(mask_for_clipboard)

# mask_for_clipboard.head(9)
stacked = formulator.loc[mask_for_clipboard, ['sheetcell', 'sheet', 'formula', 'next_cells']]

# unstack next_cells (put each in a separate row)
new_for_dedup = []
for i in stacked.index:
  for k in stacked.loc[i, 'next_cells']:
    # type(k) #str
    if '!' in k:
      # take them as they are
      new_for_dedup.append(
        [stacked.loc[i, 'sheetcell'], stacked.loc[i, 'formula'], k]
        )
      # new_for_dedup.append(k)
    else:
      new_for_dedup.append([
        stacked.loc[i, 'sheetcell']
        , stacked.loc[i, 'formula']
        ,  stacked.loc[i, 'sheet'] + '!' + k])
    
# to DataFrame
new_for_dedup = pd.DataFrame(new_for_dedup, columns=['sheetcell', 'formula', 'influencer_cell'])
# new_for_dedup = pd.DataFrame(new_for_dedup, columns=['cell'])

#todo troubleshooting e!V$28 still in til here

# remove dollar sign
# also already create influencer_sheetcell for later alteration
new_for_dedup.loc[:, 'influencer_sheetcell'] = \
  new_for_dedup.loc[:, 'influencer_cell'].str.replace(r'\$', '', regex=True)
new_for_dedup.loc[:, 'influencer_cell'] = \
  new_for_dedup.loc[:, 'influencer_cell'].str.replace(r'\$', '', regex=True)
new_for_dedup.loc[:, 'formula'] = \
  new_for_dedup.loc[:, 'formula'].str.replace(r'\$', '', regex=True)


# refurnish influencer_sheetcell with sheet to justify the 'sheet' in its name
mask_no_sheet = ~new_for_dedup.loc[:,'influencer_sheetcell'].str.contains(r'\!')
sum(mask_no_sheet)
new_for_dedup.loc[mask_no_sheet, 'influencer_sheetcell'] =\
  new_for_dedup.loc[:, 'sheetcell'].str.extract(r'^(\w{1,})\!', expand=False) \
  + '!' + new_for_dedup.loc[mask_no_sheet, 'influencer_sheetcell']
  
# generate one column without sheet prefix for use in openpyxl
new_for_dedup.loc[:, 'influencer_cell'] = \
  new_for_dedup.loc[:, 'influencer_sheetcell'].str.replace(r'^(\w{1,})\!', '', regex=True)

# nice to have: expand ranges
# mask_ranges = new_for_dedup.loc[:,'cell'].str.contains(r'\:', na=False)
# # sum(mask_ranges)
# new_for_dedup.loc[mask_ranges, 'cell']

#```



#```{python  staff formulas with sheet prefix}

#done this way because replacing parts of a regex match seems difficult:
# https://stackoverflow.com/questions/4489074/python-regular-expression-replacing-part-of-a-matched-string

#generate column sheetformula upfront, in case it's needed
new_for_dedup.loc[:, 'sheetformula'] = new_for_dedup.loc[:, 'formula']

# insert sheet prefix
unique_sheetcells = new_for_dedup.loc[:, 'sheetcell'].drop_duplicates()
# type(unique_sheetcells) #pd Series
# unique_sheetcells.shape
for u in unique_sheetcells:
  # print('u=',u)
  mask_u = new_for_dedup.loc[:, 'sheetcell'] == u
  for h in new_for_dedup.loc[mask_u, :].index:
    # print('h=',h)
    if new_for_dedup.loc[h, 'influencer_sheetcell'] in new_for_dedup.loc[h, 'sheetformula']:
      1+1
      #do nothing
    else:
      new_for_dedup.loc[h, 'sheetformula'] = \
      new_for_dedup.loc[h, 'sheetformula'].replace(
        new_for_dedup.loc[h, 'influencer_cell'],
        new_for_dedup.loc[h, 'influencer_sheetcell'],
        )
      new_for_dedup.loc[mask_u, 'sheetformula'] = new_for_dedup.loc[h, 'sheetformula']

# r.View(new_for_dedup, 'new_for_dedup')


#```


#```{python  merge to complete the from_step in the dedup chunk above}
#unique
sheetformulas_for_merge = new_for_dedup.drop_duplicates('sheetcell')

#new dna
new_dna = pure_dedup.merge(
  sheetformulas_for_merge.loc[:,['sheetcell', 'sheetformula']]
  , on='sheetcell', how='left'
  )
  
# new_dna.value_counts('sheetcell', sort=True)

# run only for from_step==1 when workbook_dna.csv is no yet present:
# old_dna = new_dna.copy()


#update dna
old_dna.update(new_dna)

updated_dna = old_dna.loc[:,['sheetcell', 'step', 'sheetformula']]

# save dna
updated_dna.to_csv('notebooks/part_results/workbook_dna.csv')

#```


#```{python  clipboard to paste next step in the dedup chunk above}


# deduplication
mask_history = dedup.loc[:,'step'] <= from_step
if len(mask_history)>0:
  print(
    'Ratio of df "dedup" currently in history: '
    , sum(mask_history)/len(mask_history))

mask_new =\
  ~new_for_dedup.loc[:,'influencer_sheetcell']\
  .isin(updated_dna.loc[mask_history,'sheetcell'])

if len(mask_new)>0:
  print(
    'Ratio of really new influencer cells compared to \
    influencer cells in the current step: '
    , sum(mask_new)/len(mask_new))

# the mentioned check in new_for_check was a step during debugging just below
new_for_check = new_for_dedup.loc[mask_new,:]

# new_for_check.to_clipboard(index=False, header=False)

new_for_check = new_for_check.drop_duplicates('influencer_sheetcell')

# prepare for clipboard
list_of_new_cells = list(new_for_check.loc[:,'influencer_sheetcell'])
# nice to have: use format:
# for_clipboard = ["step0{} = pd.DataFrame(\{'sheetcell': [".format(from_step+1)]
for_clipboard = ["stepXX = pd.DataFrame({'sheetcell': ["]
for k in list_of_new_cells:
  # type(k) #str
  for_clipboard.append("  '"+k+"',")
for_clipboard.append("  ]}) # o todo remove last comma")
for_clipboard.append("stepXX.loc[:, 'step'] = XX")
for_clipboard.append("list_of_steps = list_of_steps + [stepXX]")

# pd.DataFrame(for_clipboard).to_clipboard(index=False, header=False)
# r.View(formulator, 'formulator')
# r.View(dna, 'dna')
# r.View(for_clipboard, 'for_clipboard')
#```

###### clipboard corner

###### get variable description

#```{python   get variable description}
#todo remove test code (2 lines):
# keeper = formulator.copy()
# formulator = keeper.copy()

repairer = formulator.copy()


formulator.loc[:, 'formula'] = formulator.loc[:, 'formula'].str.replace(r'\$', '', regex=True)
formulator.loc[:, 'sheetformula'] = formulator.loc[:, 'sheetformula'].str.replace(r'\$', '', regex=True,)

#repair missing constant values in 'formula'
mask_formula_na = formulator.loc[:,'formula'].isna()
sum(mask_formula_na)
formulator.loc[mask_formula_na,'formula'] = repairer.loc[mask_formula_na,'formula']

manual_var_descriptions = {
  'u!B11': 'Rohr-Dämmung Stärke, Unterverteilung erdverlegt, m'
  ,'h!G16': 'Leitwert Rohr-Dämmung, Hauptverteilung erdverlegt, W/m/K'
  ,'h!D16': 'Rohr-Dämmung Stärke, Hauptverteilung erdverlegt, m'
  ,'s!Q32': 'Fläche Summe, m2'
  ,'s!O35': 'Fläche pro Wohngebäude, m2'
  ,'s!O36': 'Grundstückslänge, m'
  ,'u!B6': 'Grundstückslänge_, m'
  ,'s!D32': 'Einwohner aus Quelle, Personen'
  ,'u!B9': 'Temperatur Vorlauf, Unterverteilung erdverlegt, °C'
  ,'s!I35': 'Bewohner/Wohngebäude'
  ,'u!B4': 'mittlere Zweiglänge, Unterverteilung erdverlegt, m'
  ,'u!B12': 'Leitwert Rohr-Dämmung, Unterverteilung erdverlegt, W/m/K'
  ,'s!F20': 'Einwohner favorisiert Ludorf, Personen'
  ,'s!F24': 'Einwohner favorisiert Röbel, Personen'
  ,'s!F12': 'Einwohner favorisiert Groß Kelle, Personen'
  ,'s!F10': 'Einwohner favorisiert Gotthun, Personen'
  ,'s!F22': 'Einwohner favorisiert Minzow, Personen'
  ,'s!F14': 'Einwohner favorisiert Leizen, Personen'
  ,'s!F7':  'Einwohner favorisiert Bütow, Personen'
  ,'s!F8':  'Einwohner favorisiert Dambeck, Personen' #todo warum fehlt das komplett
  ,'s!F2':  'Einwohner favorisiert Bollewick, Personen'
  ,'t!E3': 'Bevölkerung, Personen'
  ,'s!G20': 'Bevölkerung Ludorf, Personen'
  ,'s!G24': 'Bevölkerung Röbel, Personen'
  ,'s!G12': 'Bevölkerung Groß Kelle, Personen'
  ,'s!G10': 'Bevölkerung Gotthun, Personen'
  ,'s!G22': 'Bevölkerung Minzow, Personen'
  ,'s!G14': 'Bevölkerung Leizen, Personen'
  ,'s!G7':  'Bevölkerung Bütow, Personen'
  ,'s!G8':'Bevölkerung Dambeck, Personen'   #todo warum fehlte das komplett
  ,'s!G2':  'Bevölkerung Bollewick, Personen'
  ,'h!D11': 'Bevölkerung Ludorf, Personen'
  ,'h!D13': 'Bevölkerung Röbel, Personen'
  ,'h!D9': 'Bevölkerung Groß Kelle, Personen'
  ,'h!D8': 'Bevölkerung Gotthun, Personen'
  ,'h!D12': 'Bevölkerung Minzow, Personen'
  ,'h!D10': 'Bevölkerung Leizen, Personen'
  ,'h!D6': 'Bevölkerung Bütow, Personen'
  ,'h!D7': 'Bevölkerung Dambeck, Personen'  #todo warum fehlte das in sheetcells
  ,'h!D5': 'Bevölkerung Bollewick, Personen'
  ,'u!B7': 'Norm-Geschwindigkeit des Wärmeträgers in Unterverteilung (UV), m/s'
  ,'u!B8': 'Geschwindigkeit des Wärmeträgers bei Auslegungsleistung in UV, m/s'
  ,'h!D17': 'Normgeschwindigkeit in Hauptverteilung, m/s'
  ,'s!O32': 'mittlere Zweiglänge, m'
  ,'s!O34': 'mittlere Zweiglänge, m'
  ,'s!I32': 'Summe Anzahl Wohngebäude'
  ,'e!D736': 'Endenergie Direktbezug aus Kollektoren, am Saisonspeicher vorbei, Jahressumme, kWh/d/Kopf'
  ,'e!C736': 'Endenergie (EE) Verbrauch  für Heizung & Warmwasser, Jahressumme, kWh/d/Kopf'
  ,'e!D737': 'Anteil Endenergie Direktbezug aus Kollektoren, am Saisonspeicher vorbei, Prozent'

  }

def get_var_description(sheetcell):
  """
  Gets in most cases the variable name for a given sheetcell from a Excel sheet like jahreslauf_roebel.xlsx
  in most cases, because mostly the desired info is placed one cell to the right from the value
  example usage: get_var_description(sheetcell='t!A42')
  """
  #remove the following test cases
  #test cases begin
  # sheetcell = 'e!A4' #example
  # sheetcell = 'e!V30' #example
  # sheetcell = 't!D99' #example
  # sheetcell = 't!DA99' #example
  # sheetcell = 'h!E21' #example
  #test cases end
  sheet = re.search(r'^(\w{1,})\!', sheetcell).groups()[0]
  col = re.search(r'^\w{1,}\!([A-Z]{1,})\d{1,}', sheetcell).groups()[0]
  row = int(re.search(r'^\w{1,}\![A-Z]{1,}(\d{1,})', sheetcell).groups()[0])
  sheetcol = re.search(r'(^\w{1,}\![A-Z]{1,})\d{1,}', sheetcell).groups()[0]
  
  
  if sheetcol in ['t!D','e!V','t!F']:
    
    delta_row_col = {
      't!D':   [[0,-2], [0,-1], [0, 1]] # 124
      ,'e!V':  [[0, 4], [0,-1], [0, 1]]	# 18
      ,'t!F':  [[0, 5], [0,-3], [0, 6]]	# 17
      }
        
    cell_1 = get_column_letter(
      column_index_from_string(col) + delta_row_col[sheetcol][0][1]
      ) + str(row + delta_row_col[sheetcol][0][0])
      
    cell_2 = get_column_letter(
      column_index_from_string(col) + delta_row_col[sheetcol][1][1]
      ) + str(row + delta_row_col[sheetcol][1][0])
      
    cell_3 = get_column_letter(
      column_index_from_string(col) + delta_row_col[sheetcol][2][1]
      ) + str(row + delta_row_col[sheetcol][2][0])
              
    result =\
    str(wbf_workbook[sheet][cell_1].internal_value or "") \
    + " " + str(wbf_workbook[sheet][cell_2].internal_value or "") \
    + ", " + str(wbf_workbook[sheet][cell_3].internal_value or "")
  
  #h
  elif (sheet=='h') & (col>='D') & (col<='S') & (row>=21) & (row<=32):
    cell_1 = col + '19' #name
    cell_2 = 'C' + str(row) #lead
    cell_3 = col + '20' #unit
      
    result = \
    str(wbf_workbook[sheet][cell_1].internal_value or "") \
    + " Leitung " \
    + str(wbf_workbook[sheet][cell_2].internal_value or "") \
    + ", " + str(wbf_workbook[sheet][cell_3].internal_value or "")
  
  #u
  elif (sheet=='u') & (col>='B') & (col<='M') & (row>=16) & (row<=18):
    cell_1 = col + '14' #name
    cell_2 = 'B' + str(row) #lead
    cell_3 = col + '15' #unit
      
    result = \
    str(wbf_workbook[sheet][cell_1].internal_value or "") \
    + " Zweigabschnitt " \
    + str(wbf_workbook[sheet][cell_2].internal_value or "") \
    + "  bei Auslegungsleistung, " + str(wbf_workbook[sheet][cell_3].internal_value or "")
    
  elif (sheet=='u') & (col>='B') & (col<='M') & (row>=23) & (row<=25):
    cell_1 = col + '21' #name
    cell_2 = 'B' + str(row) #lead
    cell_3 = col + '22' #unit
      
    result = \
    str(wbf_workbook[sheet][cell_1].internal_value or "") \
    + " Zweigabschnitt " \
    + str(wbf_workbook[sheet][cell_2].internal_value or "") \
    + "  bei Normleistung, " + str(wbf_workbook[sheet][cell_3].internal_value or "")
    
  #e  
  elif (sheet=='e') & (col>='A') & (col<='R') & (row>=4) & (row<=734):
    cell_1 = col + '2' #name
    # cell_2 = 'B' + str(row) #lead
    cell_3 = col + '3' #unit
      
    result = \
    str(wbf_workbook[sheet][cell_1].internal_value or "") \
    + ", " + str(wbf_workbook[sheet][cell_3].internal_value or "")
  
  elif sheetcell in manual_var_descriptions.keys():
    result = manual_var_descriptions[sheetcell]
    
  else:
    result = 'tbd'
  
  return re.sub(r'\s+', ' ', result).strip()

new_for_dedup['var_description'] = new_for_dedup.loc[:, 'sheetcell'].apply(get_var_description)
new_for_dedup['var_name'] = new_for_dedup.loc[:, 'var_description'].apply(make_var_name)

formulator['var_description'] = formulator.loc[:, 'sheetcell'].apply(get_var_description)
tempcol = formulator.pop('var_description'); formulator.insert(3, 'var_description', tempcol)
formulator.insert(4, 'var_name', tempcol.apply(make_var_name))

# tools to fix the remaining cells without easily found name
# paste their descriptions into the dict 'manual_var_descriptions' above
# r.View(formulator, 'formulator')
# v = formulator.loc[formulator.loc[:,'var_description']=='tbd',:'var_description']; v; len(v)
# formulator.loc[formulator.loc[:,'var_description']=='tbd', ['sheetcell']].to_clipboard(header=False, index=False)

# save formulator
# formulator.to_csv('part_results/formulator.csv')
#```


###### formula factory

#```{python  formula factory}

# baseline = pd.read_csv('part_results/formulator.csv', index_col = [0])
baseline = formulator.copy()


#repair na in sheetformula
df = baseline.copy()
# df.dtypes
mask_next_cells_na = df.loc[:,'next_cells'].isna()
sum(mask_next_cells_na)
df.loc[mask_next_cells_na,'sheetformula'] =\
  '=' + df.loc[mask_next_cells_na,'formula'].astype('str')
# r.View(df, 'df')

# df.info()

#input_values
mask_input = df.loc[:,'next_cells'].isna()
sum(mask_input)
input_values = df.loc[mask_input, :]
input_values.loc[:, ['var_name']].duplicated().sum() #0: OK
input_dict = input_values.loc[:, ['var_name', 'formula']].set_index('var_name')\
  .T.to_dict('records')[0]
#manual additions:
input_dict['Flaeche_Summe__m2']=8110448
pickle.dump(input_dict, open('input_dict_roebel.p', 'wb'))

# test=pickle.load(open('input_dict_roebel.p', 'rb')) ok

#todo remove test code line and 2 constraints in the repl_dict loop:
# df = df.drop('wordformula', axis=1).columns

# replace - dictionary 
df.loc[:, ['sheetcell']].duplicated().sum() #ok
# sort for disambiguation
sortiert = df.copy()
sortiert['row_num'] = sortiert.loc[:,'row'].astype('int')
#dont forget writing back
sortiert = sortiert.sort_values(by='row_num', ascending=False)

repl_dict = \
  sortiert.loc[:, ['sheetcell', 'var_name']].set_index('sheetcell')\
  .T.to_dict('records')[0] #solution to_dict from DataFrame
repl_dict['PI()'] = 'np.pi'
repl_dict['='] = '=\\' + '\n  '
repl_dict['LN('] = 'np.log('
repl_dict['EXP('] = 'np.exp('
repl_dict['SIN('] = 'np.sin('
repl_dict['^'] = '**'

#wordformula
df.insert(5, 'wordformula', df.loc[:, 'sheetformula'])
for i in df.index:
  for key in repl_dict.keys():
    df.loc[i, 'wordformula'] = \
    df.loc[i, 'wordformula'].replace(key, repl_dict[key] + '\\' + '\n  ')
    # df.loc[i, 'wordformula'] = \
    # df.loc[i, 'wordformula'].replace(key, repl_dict[key] + '\\' + '\n  ')
# df.loc[:10,['var_name', 'wordformula']].to_clipboard(header=False, index=False)
df.loc[:,'wordformula'] = df.loc[:,'wordformula'].str.replace(r'\\\s+$', '', regex=True)
# r.View(df, 'df')

#save df
df.to_csv('part_results/df_with_wordformula.csv')

#```


###### build calculation
#```{python}
#bis zu welcher Zeile in 'geordnet':
limit = 400


constructor = pd.read_csv('notebooks/part_results/df_with_wordformula.csv')
#remove dollar sign
constructor['next_cells'] = constructor['next_cells'].str.replace('$', '')
#only what is not an input_value
mask_not_input = constructor.loc[:,'var_name'].isin(input_dict.keys())
sum(mask_not_input) #116
constructor = constructor.loc[~mask_not_input, :]
constructor = constructor.reset_index(drop=True)

# use the fruits of the first manual approach
#try to figure out the execution order
#run only export once, then change order and reimport
# ordner = constructor.loc[:,['sheetcell', 'sheetformula', 'var_name']]
# ordner.to_csv('part_results/ordner.csv')
#
ordner = pd.read_csv('notebooks/part_results/ordner.csv').loc[:,['sheetcell']]
geordnet = ordner.merge(constructor, on='sheetcell', how='left')
#
mask_no_var_name = geordnet['var_name'].isna()
sum(mask_no_var_name)
geordnet = geordnet.loc[~mask_no_var_name, :]
geordnet = geordnet.reset_index(drop=True)

# nice to have: renaming
# geordnet.columns = list(geordnet.columns).replace('Unnamed: 0', 'old_index')
# then inactivate the following
geordnet = geordnet.drop(['Unnamed: 0'], axis=1)


# # remove duplicated var_names?
# # e.g. Bevoelkerung_Roebel__Personen
# sum(geordnet.duplicated('var_name')) #11
# # see what to keep
# # View(geordnet.loc[geordnet.duplicated('var_name', keep=False), :])
# geordnet = geordnet.drop_duplicates('var_name', keep='first')
# sum(geordnet.duplicated('var_name')) #0 ok


# #corrections_to_geordnet
geordnet.loc[geordnet.loc[:,'var_name']=='Einwohner_aus_Quelle__Personen', 'wordformula']=\
  '=\\\n\
  \\\n\
    Bevoelkerung_Gross_Kelle__Personen\\\n\
    + Bevoelkerung_Ludorf__Personen\\\n\
    + Bevoelkerung_Roebel__Personen\\\n\
    + Bevoelkerung_Gotthun__Personen\\\n\
    + Bevoelkerung_Minzow__Personen\\\n\
    + Bevoelkerung_Leizen__Personen\\\n\
    + Bevoelkerung_Buetow__Personen\\\n\
    + Bevoelkerung_Dambeck__Personen\\\n\
    + Bevoelkerung_Bollewick__Personen'
geordnet.loc[geordnet.loc[:,'var_name']=='Flaeche_pro_Wohngebaeude__m2', 'wordformula']=\
  '=\\\n\
  \\\n\
    Flaeche_Summe__m2\\\n\
    /Summe_Anzahl_Wohngebaeude'
geordnet.loc[geordnet.loc[:,'var_name']=='Bewohner_pro_Wohngebaeude', 'wordformula']=\
  '=\\\n\
  \\\n\
    Einwohner_aus_Quelle__Personen\\\n\
    /Summe_Anzahl_Wohngebaeude'
    
geordnet.loc[geordnet.loc[:,'var_name']==\
  'Jahreswaermeverluste_Zweigabschnitt_gesamt_bei_Normleistung__kWh_pro_a', 'wordformula']=\
  '=\\\n\
  \\\n\
    Jahreswaermeverluste_Zweigabschnitt_1_bei_Normleistung__kWh_pro_a\\\n\
    +Jahreswaermeverluste_Zweigabschnitt_2_bei_Normleistung__kWh_pro_a'
    
geordnet.loc[geordnet.loc[:,'var_name']==\
  'Waermeverluste_Leitung___kWh_pro_a', 'wordformula']=\
  '=\\\n\
  \\\n\
    Waermeverluste_Leitung_1__kWh_pro_a\\\n\
    +Waermeverluste_Leitung_2__kWh_pro_a\\\n\
    +Waermeverluste_Leitung_3__kWh_pro_a\\\n\
    +Waermeverluste_Leitung_4__kWh_pro_a\\\n\
    +Waermeverluste_Leitung_5__kWh_pro_a\\\n\
    +Waermeverluste_Leitung_6__kWh_pro_a\\\n\
    +Waermeverluste_Leitung_7__kWh_pro_a\\\n\
    +Waermeverluste_Leitung_8__kWh_pro_a\\\n\
    +Waermeverluste_Leitung_9__kWh_pro_a\\\n\
    +Waermeverluste_Leitung_10__kWh_pro_a\\\n\
    +Waermeverluste_Leitung_11__kWh_pro_a'
    
geordnet.loc[geordnet.loc[:,'var_name']==\
  'Gesamtverlust_Saisonspeicher__kWh_pro_a', 'wordformula']=\
  '=\\\n\
  \\\n\
    Saisonspeicherverluste_nach_oben__kWh_pro_a\\\n\
    +Saisonspeicherverluste_durch_Waermeleitung_im_Boden__kWh_pro_a\\\n\
    +Saisonspeicherverluste_durch_Niederschlag_in_der_Umgebung__kWh_pro_a\\\n\
    +Saisonspeicherverluste_durch_regionalen_Grundwasserfluss_in_Speicherumgebung__kWh_pro_a\\\n\
    +Saisonspeicherverluste_durch_Grundwasserkonvektion_in_der_Speicherumgebung__kWh_pro_a'
    
geordnet.loc[geordnet.loc[:,'var_name']==\
  'Gesamtverluste_durch_Speicherung_und_Verteilung__kWh_pro_a', 'wordformula']=\
  '=\\\n\
  \\\n\
    Jahreswaermeverlust_durch_Speicherung__kWh_pro_a\\\n\
    +Jahreswaermeverlust_im_Fernwaermenetz__kWh_pro_a'

geordnet.loc[geordnet.loc[:,'var_name']==\
  'Leistungsaufnahme_der_Pumpen__Vor__und_Ruecklauf_Zweigabschnitt_gesamt_bei_Normleistung__kW', 'wordformula']=\
  '=\\\n\
  \\\n\
    Leistungsaufnahme_der_Pumpen__Vor__und_Ruecklauf_Zweigabschnitt_1_bei_Normleistung__kW\\\n\
    +Leistungsaufnahme_der_Pumpen__Vor__und_Ruecklauf_Zweigabschnitt_2_bei_Normleistung__kW'

geordnet.loc[geordnet.loc[:,'var_name']==\
  'Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung___kW', 'wordformula']=\
  '=\\\n\
  \\\n\
    Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_1__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_2__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_3__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_4__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_5__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_6__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_7__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_8__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_9__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_10__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Normalleistung__Vor__und_Ruecklauf_Leitung_11__kW'

geordnet.loc[geordnet.loc[:,'var_name']==\
  'Leistungsaufnahme_der_Pumpen__Vor__und_Ruecklauf_Zweigabschnitt_gesamt_bei_Auslegungsleistung__kW', 'wordformula']=\
  '=\\\n\
  \\\n\
    Leistungsaufnahme_der_Pumpen__Vor__und_Ruecklauf_Zweigabschnitt_1_bei_Auslegungsleistung__kW\\\n\
    +Leistungsaufnahme_der_Pumpen__Vor__und_Ruecklauf_Zweigabschnitt_2_bei_Auslegungsleistung__kW'


geordnet.loc[geordnet.loc[:,'var_name']==\
  'Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung___kW', 'wordformula']=\
  '=\\\n\
  \\\n\
    Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_1__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_2__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_3__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_4__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_5__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_6__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_7__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_8__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_9__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_10__kW\\\n\
    +Leistungsaufnahme_der_Pumpen_bei_Auslegungsleistung__Vor__und_Ruecklauf_Leitung_11__kW'

geordnet.loc[geordnet.loc[:,'var_name']==\
  'Trassenlaenge_Leitung___m', 'wordformula']=\
  '=\\\n\
  \\\n\
    Trassenlaenge_Leitung_1__m\\\n\
    +Trassenlaenge_Leitung_2__m\\\n\
    +Trassenlaenge_Leitung_3__m\\\n\
    +Trassenlaenge_Leitung_4__m\\\n\
    +Trassenlaenge_Leitung_5__m\\\n\
    +Trassenlaenge_Leitung_6__m\\\n\
    +Trassenlaenge_Leitung_7__m\\\n\
    +Trassenlaenge_Leitung_8__m\\\n\
    +Trassenlaenge_Leitung_9__m\\\n\
    +Trassenlaenge_Leitung_10__m\\\n\
    +Trassenlaenge_Leitung_11__m'


geordnet.loc[geordnet.loc[:,'var_name']==\
  'Endenergie_Direktbezug_aus_Kollektoren__am_Saisonspeicher_vorbei__Jahressumme__kWh_pro_d_pro_Kopf', 'wordformula']=\
  '=\\\n\
  \\\n\
    sum(kurven["Endenergie_Direktbezug_aus_Kollektoren__am_Saisonspeicher_vorbei"])/2'

geordnet.loc[geordnet.loc[:,'var_name']==\
  'Endenergie__EE__Verbrauch_fuer_Heizung_u_Warmwasser__Jahressumme__kWh_pro_d_pro_Kopf', 'wordformula']=\
  '=\\\n\
  \\\n\
    sum(kurven["Endenergie__EE__Verbrauch__fuer_Heizung_u_Warmwasser"])/2'


geordnet.loc[geordnet.loc[:,'var_name']==\
  'Investition_Saisonspeicher__gesamt___Euro_', 'wordformula']=\
  '=\\\n\
  \\\n\
    Investition_Saisonspeicher__Abdeckung___Euro_\\\n\
    +Investition_Saisonspeicher__Bohrungen___Euro_\\\n\
    +Investition_Saisonspeicher__Schlitzwand___Euro_\\\n\
    +Investition_Saisonspeicher__zwei_Pufferspeicher___Euro_\\\n\
    +Investition_Technikgebaeude_am_Speicherrand__geschaetzt___Euro_'



geordnet.loc[geordnet.loc[:,'var_name']==\
  'Investition_fuer_die_gesamte_Anlage___Euro__pro_a_pro_Kopf', 'wordformula']=\
  '=\\\n\
  \\\n\
    Investitionskosten_Saisonspeicher__gesamt___Euro__pro_a_pro_Kopf\\\n\
    +Investitionskosten_Umwaelzpumpen_Unterverteilung___Euro__pro_a_pro_Kopf\\\n\
    +Investitionskosten_Umwaelzpumpen_Hauptverteilung___Euro__pro_a_pro_Kopf\\\n\
    +Investitionskosten_Kollektoren___Euro__pro_a_pro_Kopf\\\n\
    +Investitionskosten_Aufstellung__Installation_der_Kollektoren___Euro__pro_a_pro_Kopf\\\n\
    +Investitionskosten_Waermespeicher_in_den_Gebaeuden___Euro__pro_a_pro_Kopf\\\n\
    +Investitionskosten_Bodenpreis_fuer_externes_Kollektorfeld___Euro__pro_a_pro_Kopf\\\n\
    +Investitionskosten_Fernwaermeleitungen_fuer_Unterverteilung_und_Hausanschluesse___Euro__pro_a_pro_Kopf\\\n\
    +Investitionskosten_Fernwaermeleitungen_fuer_Hauptverteilung___Euro__pro_a_pro_Kopf\\\n\
    +Investitionskosten_BHKW__ohne_Energiekosten____Euro__pro_a_pro_Kopf'


geordnet.loc[geordnet.loc[:,'var_name']==\
  'laufende_Kosten___Euro__pro_a_pro_Kopf', 'wordformula']=\
  '=\\\n\
  \\\n\
    laufende_Kosten_E_Antrieb_der_Pumpen__Stromkosten___Euro__pro_a_pro_Kopf\\\n\
    +laufende_Kosten_Energiekosten_BHKW___Euro__pro_a_pro_Kopf\\\n\
    +laufende_Kosten_Ertrag_BHKW__Elektroenergie___Euro__pro_a_pro_Kopf\\\n\
    +laufende_Kosten_Waermetauscher___Euro__pro_a_pro_Kopf\\\n\
    +laufende_Kosten_Betrieb__Wartung___Euro__pro_a_pro_Kopf'


geordnet.loc[geordnet.loc[:,'var_name']==\
  'Gebaeudeanzahl_bis_Abschnittende_Zweigabschnitt_1_bei_Auslegungsleistung_', 'wordformula']=\
  '=\\\n\
  \\\n\
    mittlere_Zweiglaenge__Unterverteilung_erdverlegt__m\\\n\
    /Grundstueckslaenge__m/2'


geordnet.loc[geordnet.loc[:,'var_name']==\
  'Endenergie_Direktbezug_aus_Kollektoren__am_Saisonspeicher_vorbei__kWh_pro_d_pro_Kopf', 'wordformula']=\
  '=\\\n\
  \\\n\
    Endenergie__EE__Verbrauch_fuer_Heizung_u_Warmwasser__kWh_pro_d_pro_Kopf \\\n\
    if (beim_Verbraucher_verfuegbare_EE_aus_inneroertlichen_Kollektoren__kWh_pro_d_pro_Kopf \\\n\
    + beim_Verbraucher_verfuegbare_EE_aus_externen_Kollektoren__kWh_pro_d_pro_Kopf) \\\n\
    > Endenergie__EE__Verbrauch_fuer_Heizung_u_Warmwasser__kWh_pro_d_pro_Kopf \\\n\
    else (beim_Verbraucher_verfuegbare_EE_aus_inneroertlichen_Kollektoren__kWh_pro_d_pro_Kopf \\\n\
    + beim_Verbraucher_verfuegbare_EE_aus_externen_Kollektoren__kWh_pro_d_pro_Kopf)'




# load input_values into Environment
input_values = pickle.load(open('input_dict_roebel.p', 'rb'))
for key in input_values.keys():
  globals()[key] = input_values[key]

# calculate (in several runs) all variables for which all inputs are given
# + record order of successful execution
geordnet.insert(1, 'var_value', None) #init

geordnet = geordnet.reset_index(drop=True)
View(geordnet)
order = []
messages = []

for i in range(5):
  for i in geordnet.index:
    try:
      if i not in order:
        # print(i)
        #calculate variable
        exec(geordnet.loc[i,'var_name'] + geordnet.loc[i, 'wordformula'] )
        # exec(geordnet.loc[14,'var_name'] + geordnet.loc[14, 'wordformula'] )
        # write variable value into data frame
        exec("geordnet.loc[i, 'var_value']" + "=" + geordnet.loc[i,'var_name'])
        # print(geordnet.loc[i,'var_name'])
        # exec('print(' + geordnet.loc[i,'var_name'] +')')
        order.append(i)
      else:
        pass
        # print(i)
    except NameError:
      pass
      # print('NameError')
    else:
      pass
      # print('other Error')
  
  print(order)
  # View(geordnet)
  message = "---> " + str(len(order)) + " of " + str(len(geordnet)) + " calculations succeeded."
  print(message)
  messages.append(message)

for i in messages:
  print(i)


# then debug by manually altering var_name and wordformula in geordnet under #corrections_to_geordnet

# # debug code:
# # find out where the gaps are
# # i.e. what was not calculated
# View(geordnet.loc[geordnet.loc[:,'var_value'].isna()])
# missing = geordnet.loc[geordnet.loc[:,'var_value'].isna()]
# erster_fehlender_Index = missing.index.to_list()[0]
# erster_fehlender_Index
# aktuell_fehlende_Variable =\
#   missing.loc[erster_fehlender_Index,'var_name']
# aktuell_fehlende_Variable
# 
# # debug_i=159
# # debug_j=14
# 
# #Grund des Fehlens
# exec(geordnet.loc[erster_fehlender_Index,'var_name'] + geordnet.loc[erster_fehlender_Index, 'wordformula'] )
# 
# # Grund des Grundes 1
# mask_current_missing = geordnet.loc[:, 'var_name'] == 'endet_bei_Meter_Zweigabschnitt_1_bei_Auslegungsleistung_'
# View(geordnet.loc[mask_current_missing, ['wordformula']])
# #copy from view, because otherwise truncated
# Grund_des_Grundes_1 = 'mittlere_Zweiglaenge__Unterverteilung_erdverlegt__m'
# 
# # Grund des Grundes 2
# mask_current_missing = geordnet.loc[:, 'var_name'] == Grund_des_Grundes_1
# View(geordnet.loc[mask_current_missing, ['var_name', 'wordformula']])
# #copy from view, because otherwise truncated
# Grund_des_Grundes_2 = 'mittlere_Zweiglaenge__m'
# 
# # # Grund des Grundes 3
# # mask_current_missing = geordnet.loc[:, 'var_name'] == Grund_des_Grundes_2
# # View(geordnet.loc[mask_current_missing, ['wordformula']])
# # #copy from view, because otherwise truncated
# # Grund_des_Grundes_3 = 'mittlere_Zweiglaenge__m'
# 
# #Grund des Fehlens
# exec(geordnet.loc[mask_current_missing,['var_name']]\
#   .to_string(header=False, index=False) + \
#   geordnet.loc[mask_current_missing,['wordformula']]\
#   .to_string(header=False, index=False) )
# 
# 
# 


#---

# dazugehoerige_Wordformula = geordnet.loc[mask_current_missing, 'wordformula']
# pd.DataFrame([str(dazugehoerige_Wordformula)]).to_clipboard(index=False, header=False)
# print(dazugehoerige_Wordformula, file='~/Downloads/printer.txt')

# for debug_i in missing.index:
#   try:
#     exec(geordnet.loc[debug_i,'var_name'] + geordnet.loc[debug_i, 'wordformula'] )
#   except Exception as error:
#     print(type(error).__name__)
#     # print(error)
#   # exec(geordnet.loc[debug_i,'var_name'])
# 
# 
# # df_order = pd.DataFrame({'ord': order}) #todo delete test code?
# View(geordnet.iloc[[61]])


# # input_from_row
# # at which cell is the first input cell calculated in the moment?
# geordnet.insert(1, 'input_from_row', None)
# # View(geordnet)
# 
# for i in geordnet.index:
#   #todo remove test code
#   # i=132 #'endet_bei_Meter_Zweigabschnitt_1_bei_Auslegungsleistung_'
#   
#   # exec('next_cells_list = '+geordnet.loc[i, 'next_cells'])
#   geordnet.loc[i,'input_from_row'] = \
#     geordnet.index[(geordnet.loc[i,'var_name'] in geordnet.loc[j,'wordformula'].tolist()).any(1)].tolist()
#     geordnet.loc[i,'var_name'] in geordnet.loc[:,'var_name'].tolist() #works but is to small
#     # geordnet.index[(geordnet['sheetcell']==next_cells_list[0]).any(1)].tolist()
# 



#build calc_string
chunk_start = '\n```{python}\n'
chunk_end = '\n```\n'
build = chunk_start
build = build + 'import numpy as np' + '\n'
build = build + 'import pickle' + '\n'
build = build + 'import pprint' + '\n'
build = build + '\n'
build = build + "input_values = pickle.load(open('input_dict_roebel.p', 'rb'))" + '\n'
build = build + 'for key in input_values.keys():' + '\n'
build = build + '  globals()[key] = input_values[key]' + '\n'
build = build + '\n'
build = build + '\n'
build = build + '#Input values:\n'
build = build + 'pprint.pprint(input_values)\n'

if len(order) == len(geordnet):
  for i in order:
    build = build + chunk_end
    build = build + chunk_start
    build = build\
      + str(geordnet.loc[i, 'var_name'])\
      + str(geordnet.loc[i, 'wordformula'])
    build = build + chunk_end
    build = build + chunk_start
    build = build + '#| echo: false\n'
    build = build\
      +'round(' + geordnet.loc[i, 'var_name'] + ', 1)'
build = build + chunk_end


myfstring = f'---\n\
title: "Wirtschaftlichkeit Saisonaler Erdwärmespeicher"\n\
format: gfm\n\
editor: source\n\
---\n\
\n\
Hier entsteht gerade die Kette der Rechnungen in Python, als deren Ergebnis die 61 EUR/a/Kopf Heizkosten erwartet werden. \n\n\
{build}'

with open("Kalkulation_raw.qmd", "w") as text_file:
    print(myfstring, file=text_file)


# r.edit_file('Kalkulation_raw.qmd',)
# df.loc[:,['var_name', 'wordformula']].to_clipboard(header=False, index=False)

# #skipped trial to remove duplicates in var_name
# #to really delete those referencing to another given sheetcell, it must be checked, that none of the references are deleted
# dup = df.loc[df.duplicated('var_name', keep=False), :]; r.View(dup, 'dup')
# dup.shape #47
# clean = df.copy()
# mask_formula_in_sheetcell = clean.loc[:,'wordformula'].str.replace('=','').isin(clean.loc[:, 'sheetcell']) 
# sum(mask_formula_in_sheetcell)
# len(dup)
# clean.insert(1, 'delete', None)
# clean.loc[mask_formula_in_sheetcell,'delete'] = 'yes'

#zeitverlauf
# kurven.info()
# kurven
# zeitlauf = kurven.loc[:, ['Tag', 'Datum']]
# zeitlauf['Datum'] = pd.to_datetime((zeitlauf['Datum'])
# 
# 
# r.View(zeitlauf, 'zeitlauf')

#```

