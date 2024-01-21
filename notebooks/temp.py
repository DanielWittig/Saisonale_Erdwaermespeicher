import pandas as pd
from openpyxl import load_workbook
import webbrowser
import re
from re import findall
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt

from openpyxl.utils import get_column_letter
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.utils.cell.html?highlight=openpyxl.utils.cell.get_column_letter#openpyxl.utils.cell.get_column_letter
def xlref(row, column, zero_indexed=True):
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)
xlref(0, 0)
xlref(0, 0)
# or
get_column_letter(2) #not zero indexed

# opposite
from openpyxl.utils import column_index_from_string
